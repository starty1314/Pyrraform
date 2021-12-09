import openpyxl as excel
from openpyxl.utils import get_column_letter
import csv
import pandas as pd
import numpy as np
import os


class DataOps:
    """
    Data Operation class
    This class contains functions that can convert data
    and the functions that export data into various formats
    """
    def __init__(self):
        pass

    def export_csv(self, list_of_dicts: list, export_path: str):
        """
        Export a list of dicts to csv format
        :param export_path: Path of where to save the output csv file
        :param list_of_dicts: Data in list of dictionaries
        :return: True if no issue, raise exception if fails on something
        """
        list_of_lists = self.convert_dicts_to_list(list_of_dicts)

        try:
            path = os.path.dirname(export_path)

            if not os.path.exists(path):
                os.mkdir(path)

            with open(export_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(list_of_lists)
        except Exception as e:
            raise e
        else:
            return True

    def export_excel(self, list_of_dicts: list, export_path: str):
        """
        Export data in excel format
        :param list_of_dicts: A list of data
        :param export_path: Export file path
        :return: N/A
        """
        list_of_lists = self.convert_dicts_to_list(list_of_dicts)

        wb = excel.Workbook()

        ws = wb.active
        ws.title = "Benchmark"
        ws.sheet_properties.tabColor = "FF9900"
        wb["Benchmark"].freeze_panes = "A2"

        for row in list_of_lists:
            ws.append(row)

        # Get the max length of each column
        header = list(list_of_dicts[0])
        df = pd.DataFrame(list_of_dicts, columns=header)
        measurer = np.vectorize(len)
        column_widths = (measurer(df.values.astype(str)).max(axis=0)).tolist()
        # print(column_widths)
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width

        wb.save(filename=export_path)

    # TODO : wait until all the statistic are finished
    def export_pdf(self, list_of_dicts: list, export_path: str):
        pass

    @staticmethod
    def export_csv_pd(list_of_dicts: list, export_path: str):
        """
        Export a list of dicts to csv format by using Pandas
        :param export_path: Path of where to save the output csv file
        :param list_of_dicts: Data in list of dictionaries
        :return: result of Dataframe.to_csv, raise exception if fails on something
        """
        try:
            header = list(list_of_dicts[0])
            df = pd.DataFrame(list_of_dicts, columns=header)
            result = df.to_csv(export_path)
        except Exception as e:
            raise e
        else:
            return result

    @staticmethod
    def convert_dicts_to_list(list_of_dicts: list):
        """
        Covert a list of dicts to a list of lists(csv format) with headers
        :param list_of_dicts: A list of dicts, dicts are in the same structure
        :return: A list of lists with header(The first list)
        """
        list_of_lists = [list(list_of_dicts[0].keys())]

        for item in list_of_dicts:
            list_of_lists.append(list(item.values()))

        return list_of_lists

    @staticmethod
    def convert_to_dataframe(list_of_dicts: list):
        """
        Convert a list of dicts to Dataframe
        :param list_of_dicts: A list of dicts
        :return: Data in dataframe format
        """
        try:
            header = list(list_of_dicts[0])
            df = pd.DataFrame(list_of_dicts, columns=header)
        except Exception as e:
            raise e
        else:
            return df
